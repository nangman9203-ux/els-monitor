"""
=============================================================================
DART 일괄신고추가서류(ELS) 수집 — 주피터 올인원 버전 v9.6
=============================================================================
v9.6: v9.5와 동일 (안전 버전 재배포)

v9.5 변경점 (삼성증권 대응):
  - strike_date 정규식 확장: "최초기준가격(결정일: YYYY년...)" 패턴 추가
  - maturity_date 정규식 확장: "만기일 (예정) [YYYY년...]" 대괄호 허용
  - extract_autocall_dates: 삼성 "1차: YYYY년 / 2차: YYYY년 ..." 패턴 추가
  - split_payoff_sections: "자동조기/만기상환" 합성 표기에서 만기상환 매칭 제외

v9.4 변경점 (유지):
  - 신규 4개 증권사 (NH/미래에셋/유안타/키움) 파싱 일반화
  - 차수 마커 ①②③/(1)(2)(3)/1)2)3) 모두 지원
  - 손익구조 발췌 정규식 보정 (자동조기 (2) 차수 충돌 방지)
  - 만기일 (예정) 같은 괄호 문구 허용
  - 월쿠폰조건이 자동조기 앞에 있는 케이스 (미래에셋/메리츠) 보정

v9.3 변경점 (유지):
  - 16개 증권사 확장, corp_code 검증
  - extract_autocall_dates: 자동조기상환평가일 표 직접 추출
  - calc_freq: strike→1차→...→만기 간격 최빈값 (사용자 정의)

v9.2 변경점 (유지):
  - 월지급 ELS 대응

v9.1 변경점 (유지):
  - 배리어 % 소수점 허용, coupon 괄호없는 표기

  ⑤ (보너스) 모집총액도 가액/금액 둘 다

필요 패키지: pip install requests openpyxl
=============================================================================
"""

# ============================================================================
# 설정값 (app.py에서 주입)
# ============================================================================

API_KEY   = None  # app.py에서 설정
DATE_FROM = None
DATE_TO   = None
OUTPUT_XLSX = "els_woori_result.xlsx"

ISSUERS = [
    # === 기존 9개 (검증 완료) ===
    "한화투자증권",
    "신한투자증권",
    "아이비케이투자증권",
    "하나증권",
    "교보증권",
    "메리츠증권",
    "신영증권",
    "한국투자증권",
    "케이비증권",
    # === 신규 7개 (1년치 우리은행 신탁 실적 확인됨, 현대차는 예비) ===
    "NH투자증권",       # 1년 115종목
    "삼성증권",         # 1년 114종목
    "미래에셋증권",     # 1년 83종목
    "대신증권",         # 1년 23종목
    "키움증권",         # 1년 21종목
    "유안타증권",       # 1년 7종목
    "현대차증권",       # 1년 0종목 (예비, 향후 발행 가능성)
]

CORP_CODES = {
    # === 기존 9개 (검증 완료) ===
    "한화투자증권":      "00148610",
    "신한투자증권":      "00138321",
    "아이비케이투자증권": "00684918",
    "하나증권":          "00113465",
    "교보증권":          "00113359",
    "메리츠증권":        "00163682",
    "신영증권":          "00136721",
    "한국투자증권":      "00160144",
    "케이비증권":        "00164876",
    # === 신규 7개 (XML+API 검증 완료) ===
    "NH투자증권":        "00120182",
    "삼성증권":          "00104856",
    "미래에셋증권":      "00111722",  # 기존 예상값 00126380 → 실제는 00111722
    "대신증권":          "00110893",
    "키움증권":          "00296290",
    "유안타증권":        "00117601",  # 기존 예상값 00137725 → 실제는 00117601
    "현대차증권":        "00137997",  # 기존 예상값 00246795 → 실제는 00137997
}

# ============================================================================
# 이하는 수정할 필요 없음
# ============================================================================

import io, os, re, time, zipfile
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DART_BASE = "https://opendart.fss.or.kr/api"
CACHE_DIR = Path("./dart_cache")
CACHE_DIR.mkdir(exist_ok=True)


# ----------------------------------------------------------------------------
# 매핑
# ----------------------------------------------------------------------------

ISSUER_SHORT = {
    "한화투자증권": "한화", "교보증권": "교보", "메리츠증권": "메리츠",
    "현대차증권": "현대차", "아이비케이투자증권": "IBK",
    "신한투자증권": "신한", "하나증권": "하나",
    "NH투자증권": "NH", "미래에셋증권": "미래에셋",
    "유안타증권": "유안타", "케이비증권": "KB",
    "한국투자증권": "한국", "신영증권": "신영",
    "키움증권": "키움", "대신증권": "대신", "삼성증권": "삼성",
}

UNDERLYING_CODE_RULES = [
    ("KOSPI2",  ["KOSPI200", "KOSPI 200", "코스피200", "코스피 200"]),
    ("SPX",     ["S&P500", "S&P 500", "SNP500", "S&P"]),
    ("SX5E",    ["EuroStoxx50", "EURO STOXX 50", "유로스탁스50", "Euro Stoxx 50", "EUROSTOXX50", "EUROSTOXX 50"]),
    ("NKY",     ["NIKKEI225", "NIKKEI 225", "닛케이225", "Nikkei 225"]),
    ("HSCEI",   ["HSCEI", "항셍중국기업", "HSCE"]),
]

CIRCLED_NUM = {
    "①": 1, "②": 2, "③": 3, "④": 4, "⑤": 5, "⑥": 6,
    "⑦": 7, "⑧": 8, "⑨": 9, "⑩": 10, "⑪": 11, "⑫": 12, "⑬": 13,
}


# ----------------------------------------------------------------------------
# 유틸
# ----------------------------------------------------------------------------

def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())

def parse_korean_date(s: str) -> Optional[date]:
    m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", s)
    return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

def normalize_payoff_text(s: str) -> str:
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\(연\s*[\d.]+\s*%\)", " ", s)
    s = re.sub(r"[\d,]+\.\d+\s*%", " ", s)
    s = re.sub(r"전자공시시스템\s*dart\.fss\.or\.kr\s*Page\s*\d+", " ", s)
    s = re.sub(r"만기\s*상환", " ", s)
    s = re.sub(r"자동\s*조기\s*상환", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def split_items(full_text: str) -> list:
    return re.split(r"\[\s*모집 또는 매출의 개요\s*\]", full_text)[1:]


def extract_amount(block: str, label_pattern: str) -> Optional[int]:
    """라벨 뒤 금액 추출. USD/KRW 둘 다 지원."""
    m = re.search(rf"{label_pattern}[\s\S]{{0,80}}?USD\s*([\d,]+)", block)
    if m:
        return int(m.group(1).replace(",", ""))
    m = re.search(rf"{label_pattern}\s*([\d,]+)\s*원", block)
    if m:
        return int(m.group(1).replace(",", ""))
    return None


# ----------------------------------------------------------------------------
# mat/freq
# ----------------------------------------------------------------------------

def months_between(d1, d2): return (d2 - d1).days / 30.4375
def round_to_3_months(months): return max(3, round(months / 3) * 3)

def format_mat(months: int) -> str:
    years = months / 12
    if years == int(years): return f"{int(years)}Y"
    return f"{years:g}Y"

def calc_mat(strike_date, maturity_date):
    if not strike_date or not maturity_date: return None
    return format_mat(round_to_3_months(months_between(strike_date, maturity_date)))

def extract_autocall_dates(block):
    """
    자동조기상환평가일 표에서만 날짜 추출.
    
    공통 패턴 (메리츠/신한/KB/한화/하나/IBK 등):
      ○ 자동조기상환평가일 및 상환금액
        차수 자동조기상환평가일 상환금액
        1차 YYYY년 MM월 DD일 ...
        2차 YYYY년 MM월 DD일 ...
        ...
      ○ 자동조기상환일 : (또는) 다. (또는) 만기상환 ...
    
    이 표 안에서만 날짜 추출하므로 월쿠폰 평가일과 분리됨.
    """
    # 1차 시도: "자동조기상환평가일 및 상환금액" 헤더로 시작하는 표
    sec = re.search(
        r"자동조기상환평가일\s*및\s*상환금액"
        r"(.*?)"
        r"(?=○\s*자동조기상환일|○\s*자동조기상환\s*확정|"
        r"다\.\s*월|다\.\s*쿠폰|다\.\s*만기|"
        r"\(5\)|$)",
        block, re.DOTALL
    )
    if sec:
        target = sec.group(1)
        dates = []
        for m in re.finditer(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", target):
            try: dates.append(date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
            except ValueError: continue
        if dates:
            return dates
    
    # 🔧 v9.5: 삼성 형식 — "중간기준가격 결정일(예정) 자동조기상환여부 결정일(예정)
    #          1차: YYYY년 MM월 DD일 2차: YYYY년 MM월 DD일 ..."
    sec_samsung = re.search(
        r"(?:중간기준가격|자동조기상환여부)\s*결정일\s*\(?예정\)?"
        r"(.*?)"
        r"(?=※|만\s*기\s*평가일|중간기준가격\s*결정일\s*행사가격|\(5\)|$)",
        block, re.DOTALL
    )
    if sec_samsung:
        target = sec_samsung.group(1)
        # "1차: YYYY년 MM월 DD일" 형식의 날짜만 추출
        dates = []
        for m in re.finditer(
            r"\d+\s*차\s*[:：]\s*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일",
            target
        ):
            try: dates.append(date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
            except ValueError: continue
        if dates:
            return dates
    
    # 2차 fallback: 기존 "나. 자동조기상환" 섹션 (구버전 형식)
    sec2 = re.search(r"나\.\s*자동조기상환.*?(?=다\.\s*만기상환|\(5\)|$)", block, re.DOTALL)
    target = sec2.group(0) if sec2 else block
    dates = []
    for m in re.finditer(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", target):
        try: dates.append(date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
        except ValueError: continue
    return dates

def calc_freq(strike_date, autocall_dates, maturity_date=None, coupon_type="step"):
    """
    조기상환 간격 계산 (사용자 정의 v9.3):
      - strike → 1차, 1차 → 2차, ..., 마지막 차수 → 만기 간격을 각각 정수 개월로 반올림
      - 모든 간격이 같으면 그 값
      - 다르면 최빈값
    
    extract_autocall_dates가 자동조기상환평가일 표만 정확히 추출하므로
    별도의 monthly 필터링 불필요.
    """
    if not strike_date or not autocall_dates: return None
    seen, cleaned = set(), []
    for d in autocall_dates:
        if d <= strike_date or d in seen: continue
        seen.add(d); cleaned.append(d)
    if not cleaned: return None
    cleaned.sort()
    
    # 만기일 안전성 체크: 자동조기 마지막 날짜보다 만기가 너무 가까우면 (1개월 미만) 제외
    # 보통 자동조기 마지막 차수 ≈ 만기일이라 중복될 수 있음
    use_maturity = False
    if maturity_date and maturity_date > cleaned[-1]:
        gap_to_maturity = months_between(cleaned[-1], maturity_date)
        if gap_to_maturity >= 1.0:  # 1개월 이상 차이나야 의미있는 만기 간격
            use_maturity = True
    
    intervals = []
    prev = strike_date
    for d in cleaned:
        intervals.append(max(1, round(months_between(prev, d))))
        prev = d
    if use_maturity:
        intervals.append(max(1, round(months_between(prev, maturity_date))))
    
    if not intervals: return None
    
    # 모두 같으면 그 값, 다르면 최빈값
    if len(set(intervals)) == 1:
        return f"{intervals[0]}M"
    return f"{Counter(intervals).most_common(1)[0][0]}M"

def calc_mat_freq(strike_date, maturity_date, autocall_dates, coupon_type="step"):
    mat = calc_mat(strike_date, maturity_date)
    freq = calc_freq(strike_date, autocall_dates, maturity_date, coupon_type)
    if mat and freq: return f"{mat}/{freq}"
    if mat: return mat
    if freq: return f"?/{freq}"
    return ""


def get_sorted_unique_autocall_dates(strike_date, autocall_dates):
    if not strike_date or not autocall_dates: return []
    seen, cleaned = set(), []
    for d in autocall_dates:
        if d <= strike_date or d in seen: continue
        seen.add(d); cleaned.append(d)
    cleaned.sort()
    return cleaned


# ----------------------------------------------------------------------------
# 손익구조 섹션 분리
# ----------------------------------------------------------------------------

def split_payoff_sections(payoff_text: str) -> dict:
    # 🔧 v9.3: "만기상환" 뒤에 차수 마커가 오는 진짜 섹션 시작 찾기
    # 차수 마커 형식: ①②③ / (1)(2)(3) / 1)2)3)
    # "만기상환의 경우와 별도로" 같은 설명 문구 스킵
    # 🔧 v9.5: "자동조기/만기상환" 같은 합성 표기 제외 (삼성)
    mat_match = None
    # 1차: "만기상환" 바로 뒤에 차수 마커가 오는 경우 (진짜 섹션 시작)
    for m in re.finditer(r"만기\s*상환", payoff_text):
        # 🔧 v9.5: 앞에 "자동조기/" 가 있으면 스킵 (삼성 케이스)
        before = payoff_text[max(0, m.start()-10):m.start()]
        if re.search(r"자동조기\s*/\s*$|자동조기상환\s*/\s*$", before):
            continue
        after = payoff_text[m.end():m.end()+10]
        # 원문자, (1), 1) 중 하나
        if re.search(r"[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬]|\(\d{1,2}\)|\d{1,2}\)", after):
            mat_match = m
            break
    # 2차 fallback: 차수 마커 없으면 마지막 "만기상환" 사용 (자동조기/는 여전히 제외)
    if not mat_match:
        for m in re.finditer(r"만기\s*상환", payoff_text):
            before = payoff_text[max(0, m.start()-10):m.start()]
            if re.search(r"자동조기\s*/\s*$|자동조기상환\s*/\s*$", before):
                continue
            mat_match = m  # 마지막 것이 남음
    
    if mat_match:
        auto_text = payoff_text[:mat_match.start()]
        maturity_text = payoff_text[mat_match.start():]
    else:
        auto_text = ""
        maturity_text = payoff_text
    
    # 🔧 v9.4: auto 섹션에서 "월수익지급/쿠폰지급조건/월쿠폰" 등이 자동조기 앞에 있으면 제거
    # (미래에셋/메리츠 월지급 등에서 월쿠폰조건이 자동조기 앞에 위치하는 케이스)
    # 진짜 자동조기상환 시작점 = "자동조기상환" 뒤 30자 이내에 차수 마커(①/(1)/1))가 오는 곳
    if auto_text:
        true_auto_start = None
        for m in re.finditer(r"자동조기상환", auto_text):
            after = auto_text[m.end():m.end()+30]
            # 차수 마커가 가까이 오는 곳이 진짜 섹션 시작
            if re.search(r"[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬]|\(\d{1,2}\)|\d{1,2}\)\s*\d", after):
                true_auto_start = m
                break
        # 진짜 자동조기 시작점을 찾았고, 그 앞에 월쿠폰 키워드가 있으면 잘라냄
        if true_auto_start:
            before = auto_text[:true_auto_start.start()]
            if re.search(
                r"월\s*수익\s*지급|쿠폰\s*지급\s*조건|월\s*쿠폰|월\s*단위\s*쿠폰",
                before
            ):
                auto_text = auto_text[true_auto_start.start():]
    
    return {
        "auto": auto_text,
        "maturity": maturity_text,
        "all": payoff_text,
    }


# ----------------------------------------------------------------------------
# ko/ki — v9.1: 배리어 % 소수점 허용
# ----------------------------------------------------------------------------

def _extract_num_from_marker(marker_text: str):
    """
    마커 텍스트에서 차수 번호 추출.
    지원 형식:
      - 원문자: ①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬
      - 괄호숫자: (1) (2) (3) ... (13)
      - 닫는괄호숫자: 1) 2) 3) ... 13)
    Returns: (num: int, sub: str)  sub는 리자드용 "-1"/"-2" 표기
    """
    # 원문자
    m = re.match(r"([①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬])\s*(?:-\s*([12]))?", marker_text)
    if m:
        return CIRCLED_NUM.get(m.group(1)), m.group(2) or ""
    # (1)(2)(3) 형식
    m = re.match(r"\((\d{1,2})\)\s*(?:-\s*([12]))?", marker_text)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 13:
            return n, m.group(2) or ""
    # 1)2)3) 형식
    m = re.match(r"(\d{1,2})\)\s*(?:-\s*([12]))?", marker_text)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 13:
            return n, m.group(2) or ""
    return None, ""


def parse_payoff_conditions(payoff_text: str) -> list:
    results = []
    # 🔧 v9.3: 차수 형식 확장 — ①②③ / (1)(2)(3) / 1)2)3) 모두 매칭
    # 원문자 우선, 그 다음 괄호숫자, 마지막 닫는괄호숫자
    marker_pat = re.compile(
        r"([①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬]\s*(?:-\s*[12])?)"      # 원문자(±리자드)
        r"|(\(\d{1,2}\)\s*(?:-\s*[12])?)"                  # (1)(2)
        r"|(?<![\d.])(\d{1,2}\)\s*(?:-\s*[12])?)"          # 1)2)  (앞에 숫자/점 없을 때)
    )
    
    raw_markers = list(marker_pat.finditer(payoff_text))
    
    # 각 매칭에서 (num, sub, start, end) 추출
    parsed_markers = []
    for m in raw_markers:
        marker_text = m.group(0)
        num, sub = _extract_num_from_marker(marker_text)
        if num is None: continue
        parsed_markers.append({
            "num": num, "sub": sub,
            "start": m.start(), "end": m.end(),
            "raw": marker_text,
        })
    
    # 중복 제거 (같은 (num, sub)는 첫 번째만)
    seen_keys = set()
    filtered = []
    for pm in parsed_markers:
        key = (pm["num"], pm["sub"])
        if key in seen_keys: continue
        seen_keys.add(key)
        filtered.append(pm)
    
    for i, pm in enumerate(filtered):
        num = pm["num"]
        sub = pm["sub"]
        start = pm["end"]
        if i + 1 < len(filtered):
            end = filtered[i+1]["start"]
        else:
            tail_m = re.search(r"[※*]", payoff_text[start:])
            end = start + tail_m.start() if tail_m else len(payoff_text)
        content = payoff_text[start:end]
        
        is_maturity = bool(re.search(r"만기평가", content))
        # 🔧 v9.3: KI 키워드 확장 (키움 "어느 하나라도/하나도 ... 하락한 적이 있")
        is_ki_touch = bool(re.search(
            r"단\s*1\s*회라도|한\s*번이라도|한\s*차례라도|"
            r"어느\s*하나(?:라도|도)[\s\S]{0,100}?하락한\s*적이\s*있",
            content
        ))
        is_loss = bool(re.search(r"\d+(?:\.\d+)?\s*%\s*(?:미만|보다\s*작)", content)) and not is_ki_touch
        
        # 🔧 v9.3: 배리어 % 정규식 확장 — 다양한 표기 지원
        # 기존: "최초기준가격의 85% 이상"
        # 신규: "[최초기준가격 × 80%]" (유안타), "최초기준가격의 [85%]" (미래에셋),
        #       "최초기준가격의 85%인 85/85 이상" (키움)
        barrier = None
        # 패턴1: 일반 "X% 이상/미만/보다 크/보다 작"
        bm = re.search(r"(\d+)(?:\.\d+)?\s*%\s*(?:이상|미만|초과|보다\s*작|보다\s*크)", content)
        if bm:
            barrier = int(bm.group(1))
        else:
            # 패턴2: "[최초기준가격 × X%]" (유안타) 또는 "[X%]" (미래에셋 대괄호)
            bm = re.search(r"\[\s*(?:최초기준가격\s*[×x]\s*)?(\d+)(?:\.\d+)?\s*%\s*\]", content)
            if bm:
                barrier = int(bm.group(1))
            else:
                # 패턴3: "최초기준가격의 X%인 ..." (키움)
                bm = re.search(r"최초기준가격의\s*(\d+)(?:\.\d+)?\s*%\s*인", content)
                if bm:
                    barrier = int(bm.group(1))
        
        return_pct = None
        ret_m = re.search(
            r"(?:액면금액|액면가액|원금|총액면금액|원\s*금)\s*[*XxＸ×x]\s*"
            r"(?:\[?\s*100\s*%?\s*\+\s*)?([\d.]+)\s*%",
            content
        )
        if ret_m:
            v = float(ret_m.group(1))
            # "원금 × 105%" 같은 경우 100을 빼야 함
            return_pct = v - 100 if v >= 100 else v
        else:
            for pm in re.finditer(r"(\d+(?:\.\d+)?)\s*%", content):
                v = float(pm.group(1))
                if barrier is not None and abs(v - barrier) < 0.01:
                    continue
                return_pct = v
                break
        
        results.append({
            "num": num, "sub": sub,
            "barrier": barrier, "return_pct": return_pct,
            "is_ki": is_ki_touch,
            "is_maturity": is_maturity,
            "is_loss": is_loss,
            "content": content[:200],
        })
    return results


def annualize_return(cumulative_pct, months):
    if cumulative_pct is None or months is None or months <= 0: return None
    return cumulative_pct * (12.0 / months)


def calc_ko_ki_v82(payoff_raw, ki_level, coupon_type, has_lizard,
                    strike_date, autocall_dates):
    if not payoff_raw: return ""
    
    text = re.sub(r"\s+", " ", payoff_raw)
    text = re.sub(r"전자공시시스템\s*dart\.fss\.or\.kr\s*Page\s*\d+", " ", text)
    text = re.sub(r"\s+", " ", text)
    
    sections = split_payoff_sections(text)
    
    auto_conds = parse_payoff_conditions(sections["auto"])
    auto_main = [c for c in auto_conds 
                 if c["sub"] in ("", "1") and not c["is_ki"]]
    auto_main.sort(key=lambda c: c["num"])
    
    mat_conds = parse_payoff_conditions(sections["maturity"])
    # is_maturity=True 만 선택 (만기 텍스트 뒤 "1)쿠폰 + 2)원금" 같은 잡음 차수 마커 제외)
    mat_main = [c for c in mat_conds
                if c["is_maturity"] and not c["is_ki"] and not c["is_loss"]]
    # 폴백: is_maturity가 모두 False인 경우 (일부 증권사 표기 다를 수 있음)
    if not mat_main:
        mat_main = [c for c in mat_conds
                    if not c["is_ki"] and not c["is_loss"]]
    mat_main.sort(key=lambda c: c["num"])
    
    main_barriers = []
    main_return_map = {}
    for c in auto_main:
        if c["barrier"] is not None:
            main_barriers.append(str(c["barrier"]))
        if c["return_pct"] is not None:
            main_return_map[c["num"]] = c["return_pct"]
    
    if mat_main:
        ko_barrier = mat_main[0]["barrier"]
        if ko_barrier is not None:
            main_barriers.append(str(ko_barrier))
    
    if not main_barriers: return ""
    schedule = "-".join(main_barriers)
    
    sorted_dates = get_sorted_unique_autocall_dates(strike_date, autocall_dates)
    lizard_parts = []
    if has_lizard and strike_date and sorted_dates:
        lizard_conds = [c for c in auto_conds if c["sub"] == "2" and not c["is_ki"]]
        lizard_conds.sort(key=lambda c: c["num"])
        
        for lc in lizard_conds:
            if lc["barrier"] is None: continue
            nth = lc["num"]
            eval_date = sorted_dates[nth - 1] if 1 <= nth <= len(sorted_dates) else None
            
            multiplier_str = "1.0x"
            if eval_date is not None:
                months_elapsed = round((eval_date - strike_date).days / 30.4375)
                base_ret = main_return_map.get(lc["num"])
                lizard_ret = lc["return_pct"]
                
                if (base_ret is not None and lizard_ret is not None 
                    and months_elapsed > 0):
                    base_ann = annualize_return(base_ret, months_elapsed)
                    lizard_ann = annualize_return(lizard_ret, months_elapsed)
                    if base_ann is not None and abs(base_ann) > 0.001:
                        ratio = lizard_ann / base_ann
                        matched = False
                        for target in [0.25, 0.333, 0.5, 0.667, 0.75, 1.0, 1.5, 2.0]:
                            if abs(ratio - target) < 0.05:
                                multiplier_str = (f"{target:.1f}x" if target == round(target, 1)
                                                  else f"{target:.2f}x")
                                matched = True
                                break
                        if not matched:
                            multiplier_str = f"{ratio:.2f}x"
            
            lizard_parts.append(f"{lc['num']}-{lc['barrier']}-{multiplier_str}")
    
    ki_part = f"/{ki_level}" if ki_level else ""
    
    monthly_barrier_part = ""
    if coupon_type == "monthly":
        # 패턴1: 신한/KB "월수익지급평가일...최초기준가격의 65% 보다 크거나"
        # 패턴2: 한화 "월 수익 지급...최초기준가격의 60% 이상"
        # 패턴3: 메리츠 "쿠폰지급 평가일...최초기준가격의 60% 이상"
        m = re.search(
            r"(?:월\s*수익\s*지급|쿠폰\s*지급|월\s*쿠폰)[\s\S]{0,300}?"
            r"최초기준가격의\s*(\d+)(?:\.\d+)?\s*%\s*(?:이상|보다\s*크거나)",
            text)
        if m: monthly_barrier_part = f"/{m.group(1)}"
    
    # 🔧 v9.2: 순서 변경 — 배리어 → KI → 리자드 → 월쿠폰
    result = schedule
    result += ki_part
    for lp in lizard_parts:
        result += f"/{lp}"
    result += monthly_barrier_part
    return result


# ----------------------------------------------------------------------------
# KI 판별
# ----------------------------------------------------------------------------

def detect_KI(payoff_raw: str) -> tuple:
    if not payoff_raw:
        return False, None
    
    text = re.sub(r"\s+", " ", payoff_raw)
    text = re.sub(r"\s+", " ", text)
    
    sections = split_payoff_sections(text)
    mat_section = sections["maturity"]
    
    # 🔧 v9.3: 차수 카운트 — 원문자/괄호숫자/닫는괄호숫자 모두
    circled = set(re.findall(r"[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬]", mat_section))
    paren_nums = set(re.findall(r"\((\d{1,2})\)", mat_section))
    close_nums = set(re.findall(r"(?<![\d.])(\d{1,2})\)", mat_section))
    # 가장 큰 카운트 사용 (혼용 방지)
    marker_count = max(len(circled), len(paren_nums), len(close_nums))
    
    # 🔧 v9.3: KI 키워드 확장
    # 기존: "단 1회라도", "한 번이라도", "한 차례라도"
    # 신규: "어느 하나라도/하나도 ... 하락한 적이 있는" (키움)
    ki_keyword = re.search(
        r"단\s*1\s*회라도|한\s*번이라도|한\s*차례라도|"
        r"어느\s*하나(?:라도|도)[\s\S]{0,100}?하락한\s*적이\s*있",
        mat_section
    )
    
    if marker_count < 3 or not ki_keyword:
        return False, None
    
    # 🔧 v9.3: KI 레벨 추출 — 다양한 표기 지원
    # 패턴1: "최초기준가격의 X% 보다 작/미만" (기존)
    ki_m = re.search(
        r"(?:단\s*1\s*회라도|한\s*번이라도|한\s*차례라도|어느\s*하나(?:라도|도))[\s\S]{0,200}?"
        r"최초기준가격의\s*(\d+)(?:\.\d+)?\s*%\s*(?:보다\s*작|미만)",
        mat_section
    )
    if ki_m:
        return True, int(ki_m.group(1))
    
    # 패턴2: 키움 형식 "X%인 X/X 미만"
    ki_m = re.search(
        r"(?:단\s*1\s*회라도|한\s*번이라도|한\s*차례라도|어느\s*하나(?:라도|도))[\s\S]{0,200}?"
        r"최초기준가격의\s*(\d+)(?:\.\d+)?\s*%\s*인",
        mat_section
    )
    if ki_m:
        return True, int(ki_m.group(1))
    
    # 패턴3: "[최초기준가격 × X%]" (대괄호 + 곱셈, 유안타 KI)
    ki_m = re.search(
        r"(?:단\s*1\s*회라도|한\s*번이라도|한\s*차례라도|어느\s*하나(?:라도|도))[\s\S]{0,200}?"
        r"\[\s*최초기준가격\s*[×x]\s*(\d+)(?:\.\d+)?\s*%\s*\]",
        mat_section
    )
    if ki_m:
        return True, int(ki_m.group(1))
    
    return True, None


# ----------------------------------------------------------------------------
# 데이터 클래스
# ----------------------------------------------------------------------------

@dataclass
class ELSItem:
    source_file: str = ""
    receipt_date: Optional[date] = None
    is_amendment: bool = False

    issuer_short: str = ""
    issuer_full: str = ""
    series_no: str = ""
    note: str = ""
    currency: str = "KRW"
    pricing_date: Optional[date] = None
    strike_date: Optional[date] = None
    structure: str = ""
    und1: str = ""
    und2: str = ""
    und3: str = ""
    mat_freq: str = ""
    ko_ki: str = ""
    price_pct: Optional[float] = None
    coupon_pct: Optional[float] = None

    full_name: str = ""
    offering_total: Optional[int] = None
    issue_price: Optional[int] = None
    face_value: Optional[int] = None
    maturity_date: Optional[date] = None
    offering_method: str = ""
    sold_via_woori: bool = False
    maturity_barrier: Optional[int] = None
    has_KI: bool = False
    KI_level: Optional[int] = None
    coupon_type: str = "step"
    has_lizard: bool = False
    underlying_raw: list = field(default_factory=list)
    autocall_dates: list = field(default_factory=list)
    fx_rate: Optional[float] = None  # USD 종목의 원화 환산 적용환율 (원/USD)


STRUCTURE_MAP = {
    (False, "step",    False): "step_down_noKI",
    (True,  "step",    False): "step_down",
    (False, "monthly", False): "stepdown_noKI_coupon",
    (True,  "monthly", False): "stepdown_KI_coupon",
    (False, "step",    True ): "stepdown_noKI_lizard",
    (True,  "step",    True ): "stepdown_KI_lizard",
    (True,  "monthly", True ): "stepdown_KI_coupon_lizard",
    (False, "monthly", True ): "stepdown_noKI_coupon_lizard",
}


# ----------------------------------------------------------------------------
# 종목 추출
# ----------------------------------------------------------------------------

def extract_underlyings(block):
    found = []
    sec = re.search(r"Ⅳ\.\s*기초자산에 관한 사항.*?(?=Ⅴ\.|$)", block, re.DOTALL)
    target = sec.group(0) if sec else block
    seen = set()
    for code, patterns in UNDERLYING_CODE_RULES:
        for p in patterns:
            if re.search(re.escape(p), target, re.IGNORECASE):
                if code not in seen: found.append(code); seen.add(code)
                break
    if not found:
        for code, patterns in UNDERLYING_CODE_RULES:
            for p in patterns:
                if re.search(re.escape(p), block, re.IGNORECASE):
                    if code not in seen: found.append(code); seen.add(code)
                    break
    return found


def extract_offering_method(block):
    for m in re.finditer(r"2\.\s*공모방법\s*(.+?)(?=3\.\s*공모가격|3\.\s*공모방법)", block, re.DOTALL):
        first_200 = m.group(0)[:200]
        if re.search(r"\.{5,}", first_200):
            continue
        return re.sub(r"\s+", " ", m.group(1)).strip()[:500]
    return ""


def extract_item(block, issuer_full):
    item = ELSItem()
    item.issuer_full = issuer_full
    item.issuer_short = ISSUER_SHORT.get(issuer_full, issuer_full)

    name_m = re.search(r"\[\s*종목명\s*:\s*([^\]]+)\]", block)
    if name_m: item.full_name = name_m.group(1).strip()
    
    if name_m:
        name_text = name_m.group(1)
        ho_m = re.search(r"(?:제\s*)?(\d+)\s*[호회]", name_text)
        if ho_m: item.series_no = ho_m.group(1)
    if not item.series_no:
        ho_m = re.search(r"제\s*(\d+)\s*[호회]", block[:500])
        if ho_m: item.series_no = ho_m.group(1)
    if not item.series_no:
        return None

    # 🔧 v9.1: 모집총액·발행가액·액면가액 — '가액'+'금액' 둘 다 매칭
    item.offering_total = extract_amount(block, r"모\s*집\s*총\s*액")
    item.issue_price = extract_amount(block, r"1\s*증\s*권\s*당\s*발\s*행\s*(?:가\s*액|금\s*액)")
    item.face_value = extract_amount(block, r"1\s*증\s*권\s*당\s*액\s*면\s*(?:가\s*액|금\s*액)")
    
    # 🔧 v9.4: "만기일(예정) YYYY년 MM월 DD일" 같은 괄호 문구 허용 (키움)
    # 🔧 v9.5: "만 기 일 (예정) [YYYY년 MM월 DD일]" 대괄호 허용 (삼성)
    m = re.search(
        r"만\s*기\s*일\s*(?:\([^)]*\))?\s*\[?\s*"
        r"(\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)",
        block
    )
    if m: item.maturity_date = parse_korean_date(m.group(1))
    
    # strike_date 추출
    # 1차: 일반 "최초기준가격평가일 YYYY년 MM월 DD일" (기존)
    m = re.search(r"최초기준가격평가일\s*[:：]?\s*(\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)", block)
    if m: item.strike_date = parse_korean_date(m.group(1))
    # 🔧 v9.5: 삼성 "최초기준가격(결정일: YYYY년 MM월 DD일)" 추가
    if not item.strike_date:
        m = re.search(
            r"최초기준가격\s*\(\s*결정일\s*[:：]\s*(\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)",
            block
        )
        if m: item.strike_date = parse_korean_date(m.group(1))
    # 🔧 v9.5: 삼성 "최초기준가격 결정일" (괄호 없는 케이스) 폴백
    if not item.strike_date:
        m = re.search(
            r"최초기준가격\s*결정일\s*\(?예정\)?\s*[:：]?\s*\[?\s*(\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)",
            block
        )
        if m: item.strike_date = parse_korean_date(m.group(1))

    head = block[:3000]
    item.currency = "USD" if re.search(r"\b달러\b|\bUSD\b", head) else "KRW"
    
    # USD 종목: 원화 환산 적용환율 추출
    # 패턴: "매매기준율 @ 1,507.20원/USD" 또는 "@ 1,507원/USD"
    if item.currency == "USD":
        fx_m = re.search(
            r"매매기준율\s*@\s*([\d,]+(?:\.\d+)?)\s*원\s*/\s*USD",
            block
        )
        if fx_m:
            try:
                item.fx_rate = float(fx_m.group(1).replace(",", ""))
            except ValueError:
                pass

    und_codes = extract_underlyings(block)
    item.underlying_raw = und_codes
    item.und1 = und_codes[0] if len(und_codes) >= 1 else ""
    item.und2 = und_codes[1] if len(und_codes) >= 2 else ""
    item.und3 = und_codes[2] if len(und_codes) >= 3 else ""

    method_text = extract_offering_method(block)
    item.offering_method = method_text
    item.sold_via_woori = ("우리은행" in method_text) and ("신탁" in method_text)

    raw_payoff = ""
    # 🔧 v9.4: NH/미래에셋의 자동조기 차수 "(2) 2차"에 걸리지 않도록
    # 진짜 다음 섹션 "(2) 예상", "(2) 만기" 같은 한글 명사로 시작하는 것만 매칭
    payoff_m = re.search(
        r"\(1\)\s*상황별 손익구조.*?"
        r"(?=\(2\)\s*(?:예상|만기|손익|기초자산|자동조기상환\s*조건|투자|기타))",
        block, re.DOTALL
    )
    if not payoff_m:
        # fallback: 기존 방식 (차수 (2)가 없는 일반 ELS)
        payoff_m = re.search(r"\(1\)\s*상황별 손익구조.*?(?=\(2\))", block, re.DOTALL)
    if payoff_m:
        raw_payoff = payoff_m.group(0)
        p = normalize_payoff_text(raw_payoff)

        # 🔧 v9.1: 배리어 정규식 소수점 허용
        mb = re.search(r"만기평가가격이\s*모두\s*(?:최초기준\s*)?(?:\S+\s+){0,3}가격의\s*(\d+)(?:\.\d+)?\s*%\s*이상", p) \
          or re.search(r"만기평가가격이\s*모두\s*최초기준가격의\s*(\d+)(?:\.\d+)?\s*%\s*이상", p)
        if mb: item.maturity_barrier = int(mb.group(1))
        
        item.has_KI, item.KI_level = detect_KI(raw_payoff)
        
        item.coupon_type = "monthly" if re.search(
            r"월\s*수익\s*지급|쿠폰\s*지급\s*조건|월\s*쿠폰|월\s*단위\s*쿠폰", p
        ) else "step"
        # 🔧 v9.3: 리자드 검출 — 원문자/괄호숫자/닫는괄호숫자 모두 지원
        lz_hits = re.findall(
            r"([①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬]|\(\d{1,2}\)|(?<![\d.])\d{1,2}\))\s*-\s*([12])", p
        )
        d = defaultdict(set)
        for ch, n in lz_hits: d[ch].add(n)
        item.has_lizard = any(s == {"1", "2"} for s in d.values())

        # 🔧 v9.1: coupon — 괄호 안/밖 둘 다 매칭
        # 우선순위 1: "(연 11.50%)" 식 (이전 방식 — 더 정확)
        coupon_hits = re.findall(r"\(\s*연\s*([\d.]+)\s*%\s*\)", raw_payoff)
        if not coupon_hits:
            # 우선순위 2: 괄호 없이 "연 15.4000%" 또는 "연 11.00% 수준"
            coupon_hits = re.findall(r"연\s*([\d.]+)\s*%", raw_payoff)
        if coupon_hits:
            most = Counter(coupon_hits).most_common(1)[0][0]
            item.coupon_pct = float(most)

    item.autocall_dates = extract_autocall_dates(block)
    item.structure = STRUCTURE_MAP.get((item.has_KI, item.coupon_type, item.has_lizard), "UNKNOWN")
    
    if item.issue_price and item.face_value:
        item.price_pct = round(item.issue_price / item.face_value * 100, 2)
    
    item.mat_freq = calc_mat_freq(item.strike_date, item.maturity_date,
                                    item.autocall_dates, item.coupon_type)
    item.ko_ki = calc_ko_ki_v82(
        raw_payoff, item.KI_level, item.coupon_type, item.has_lizard,
        item.strike_date, item.autocall_dates
    )
    return item


# ----------------------------------------------------------------------------
# 기재정정 중복 제거
# ----------------------------------------------------------------------------

def dedup_amendments(items):
    by_key = defaultdict(list)
    for it in items:
        key = (it.issuer_full, it.series_no)
        by_key[key].append(it)
    result = []
    dup_count = 0
    for (issuer, series_no), group in by_key.items():
        if len(group) == 1:
            result.append(group[0]); continue
        dup_count += 1
        group.sort(key=lambda x: (x.receipt_date or date.min, x.is_amendment), reverse=True)
        chosen = group[0]
        if chosen.is_amendment or any(x.is_amendment for x in group):
            chosen.note = "기재정정"
        result.append(chosen)
        print(f"      [중복] {issuer} 제{series_no}호 ({len(group)}건):")
        for x in group:
            marker = "✅선택" if x is chosen else "  제외"
            amend = "[정정]" if x.is_amendment else "[원본]"
            print(f"        {marker} {amend} {x.source_file} 접수={x.receipt_date}")
    print(f"      → 중복 {dup_count}건 정리 완료")
    return result


# ----------------------------------------------------------------------------
# DART API
# ----------------------------------------------------------------------------

def resolve_corp_code(api_key, company_name):
    if company_name in CORP_CODES:
        return CORP_CODES[company_name]
    cache = CACHE_DIR / "corpCode.xml"
    if not cache.exists():
        print("[corpCode] 전체 회사 리스트 다운로드 중... (최초 1회)")
        r = requests.get(f"{DART_BASE}/corpCode.xml", params={"crtfc_key": api_key}, timeout=30)
        r.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            cache.write_bytes(z.read(z.namelist()[0]))
    xml = cache.read_text(encoding="utf-8")
    listed, fallback = None, None
    for m in re.finditer(r"<list>(.*?)</list>", xml, re.DOTALL):
        blk = m.group(1)
        name_m = re.search(r"<corp_name>([^<]+)</corp_name>", blk)
        if not name_m or name_m.group(1).strip() != company_name: continue
        code = re.search(r"<corp_code>(\d+)</corp_code>", blk).group(1)
        stock_m = re.search(r"<stock_code>\s*([^<\s]*)\s*</stock_code>", blk)
        stock = stock_m.group(1) if stock_m else ""
        if stock: listed = code; break
        else: fallback = fallback or code
    if listed: return listed
    if fallback: return fallback
    raise RuntimeError(f"회사명 '{company_name}' 을 찾을 수 없습니다.")


def list_disclosures(api_key, corp_code, bgn_de, end_de, pblntf_ty="C"):
    results = []
    page = 1
    while True:
        params = dict(crtfc_key=api_key, corp_code=corp_code, bgn_de=bgn_de, end_de=end_de,
                      pblntf_ty=pblntf_ty, page_no=page, page_count=100)
        r = requests.get(f"{DART_BASE}/list.json", params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        if data.get("status") == "013": break
        if data.get("status") != "000":
            raise RuntimeError(f"DART API 에러: {data.get('message')}")
        results.extend(data.get("list", []))
        if page >= data.get("total_page", 1): break
        page += 1
        time.sleep(0.1)
    return results


def filter_target_reports(items):
    return [it for it in items
            if "일괄신고추가서류" in it.get("report_nm", "")
            and "파생결합증권" in it.get("report_nm", "")
            and "주가연계증권" in it.get("report_nm", "")]


def download_document(api_key, rcept_no):
    cache = CACHE_DIR / f"{rcept_no}.zip"
    if cache.exists(): return cache.read_bytes()
    r = requests.get(f"{DART_BASE}/document.xml",
                     params={"crtfc_key": api_key, "rcept_no": rcept_no}, timeout=60)
    r.raise_for_status()
    cache.write_bytes(r.content)
    time.sleep(0.2)
    return r.content


def xml_zip_to_text(zip_bytes):
    parts = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in z.namelist():
            if not name.lower().endswith(".xml"): continue
            raw = z.read(name)
            try: s = raw.decode("utf-8")
            except UnicodeDecodeError: s = raw.decode("cp949", errors="ignore")
            parts.append(s)
    text = "\n".join(parts)
    text = re.sub(r"<(script|style)\b[^>]*>.*?</\1>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"</(?:p|div|tr|br|li|h[1-6]|table)\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = (text.replace("&nbsp;", " ").replace("&amp;", "&")
                .replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"'))
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n\n", text)
    return text


# ----------------------------------------------------------------------------
# 엑셀 출력
# ----------------------------------------------------------------------------

EXCEL_COLUMNS = [
    ("발행사",       "issuer_short",   10),
    ("호수",         "series_no",       8),
    ("비고",         "note",           12),
    ("Cur",          "currency",        6),
    ("Pricing date", "pricing_date",   12),
    ("Strike date",  "strike_date",    12),
    ("structure",    "structure",      22),
    ("und1",         "und1",            8),
    ("und2",         "und2",            8),
    ("und3",         "und3",            8),
    ("mat/freq",     "mat_freq",       10),
    ("ko/ki",        "ko_ki",          40),
    ("price",        "price_pct",       8),
    ("coupon",       "coupon_pct",      8),
]


def export_to_excel(items, output_path):
    targets = [it for it in items if it.sold_via_woori]
    # 정렬: 통화(KRW 먼저, USD 나중) → 접수일 → 발행사 → 호수
    targets.sort(key=lambda x: (
        0 if x.currency == "KRW" else 1,
        x.receipt_date or date.min,
        x.issuer_short,
        int(x.series_no or 0)
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = "은행신탁ELS"

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [c[0] for c in EXCEL_COLUMNS]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, border

    body_font = Font(name="맑은 고딕", size=10)
    body_align_center = Alignment(horizontal="center", vertical="center")
    for r, it in enumerate(targets, start=2):
        for c_idx, (_, attr, _) in enumerate(EXCEL_COLUMNS, 1):
            val = getattr(it, attr, None)
            if val is None or val == "": val = None
            cell = ws.cell(r, c_idx, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = body_align_center
            if attr == "series_no":
                if val is not None: cell.value = int(val)
                cell.number_format = "0"
            elif attr in ("pricing_date", "strike_date"):
                cell.number_format = "yyyy-mm-dd"
            elif attr == "price_pct":
                cell.number_format = '0.00"%"'
            elif attr == "coupon_pct":
                cell.number_format = '0.00"%"'

    for idx, (_, _, width) in enumerate(EXCEL_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(str(output_path))
    print(f"\n💾 Excel 저장 완료: {output_path}  ({len(targets)}종목)")


# ----------------------------------------------------------------------------
# 증권사 1개 처리
# ----------------------------------------------------------------------------

def process_issuer(api_key, issuer_full, date_from, date_to):
    print(f"\n{'='*60}")
    print(f"◆ {issuer_full}")
    print('='*60)
    
    try:
        corp_code = resolve_corp_code(api_key, issuer_full)
        print(f"  corp_code: {corp_code}")
    except Exception as e:
        print(f"  ❌ corp_code 조회 실패: {e}")
        return []
    
    bgn, end = date_from.replace("-", ""), date_to.replace("-", "")
    listings = list_disclosures(api_key, corp_code, bgn, end)
    targets = filter_target_reports(listings)
    amend_cnt = sum(1 for t in targets if "기재정정" in t.get("report_nm", ""))
    print(f"  ELS 일괄신고추가서류: {len(targets)}건 (기재정정 {amend_cnt}건)")
    
    if not targets: return []
    
    all_items = []
    for i, t in enumerate(targets, 1):
        rcept_no, rcept_dt = t["rcept_no"], t["rcept_dt"]
        is_amend = "기재정정" in t.get("report_nm", "")
        tag = "[정정]" if is_amend else "     "
        try:
            zbytes = download_document(api_key, rcept_no)
            text = xml_zip_to_text(zbytes)
            items = []
            for b in split_items(text):
                it = extract_item(b, issuer_full)
                if it is None: continue
                it.source_file = f"{rcept_no}.xml"
                it.receipt_date = datetime.strptime(rcept_dt, "%Y%m%d").date()
                it.pricing_date = monday_of_week(it.receipt_date)
                it.is_amendment = is_amend
                items.append(it)
            all_items.extend(items)
            woori_cnt = sum(1 for x in items if x.sold_via_woori)
            print(f"  ({i}/{len(targets)}) {tag} {rcept_no}: {len(items)}종목 (우리 {woori_cnt})")
        except Exception as e:
            print(f"  ({i}/{len(targets)}) {tag} {rcept_no}: 실패 {e}")
    
    return all_items


# ----------------------------------------------------------------------------
# 메인 실행
# ----------------------------------------------------------------------------

def run(api_key, issuers, date_from, date_to, output_xlsx):
    print(f"\n{'#'*60}")
    print(f"# DART ELS 수집 v9.6 (16개 증권사, 삼성 대응)")
    print(f"# 기간: {date_from} ~ {date_to}")
    print(f"# 증권사: {', '.join(issuers)}")
    print('#'*60)
    
    all_items = []
    for issuer in issuers:
        items = process_issuer(api_key, issuer, date_from, date_to)
        all_items.extend(items)
    
    print(f"\n{'='*60}")
    print(f"◆ 전체 종합")
    print('='*60)
    print(f"  총 종목: {len(all_items)}")
    
    print(f"\n[중복 제거]")
    all_items = dedup_amendments(all_items)
    print(f"  제거 후: {len(all_items)}")
    
    print(f"\n[우리은행 신탁 필터]")
    woori = [x for x in all_items if x.sold_via_woori]
    print(f"  통과: {len(woori)}종목")
    by_issuer = Counter(x.issuer_short for x in woori)
    for issuer, cnt in by_issuer.most_common():
        print(f"    - {issuer}: {cnt}종목")
    
    print(f"\n[Structure 분포]")
    structures = Counter(x.structure for x in woori)
    for s, cnt in structures.most_common():
        print(f"    - {s}: {cnt}종목")
    
    print(f"\n[엑셀 저장]")
    export_to_excel(all_items, output_xlsx)
    return all_items


# ============================================================================
# 모듈로 사용 시 직접 실행하지 않음
# ============================================================================
# 이 파일은 app.py에서 import하여 사용합니다.
# 단독 실행하려면 주피터에서 v9.6 원본을 사용하세요.
